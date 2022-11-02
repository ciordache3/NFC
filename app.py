from flask import Flask, render_template, redirect, request
import openpyxl

app=Flask(__name__)
app.secret_key="1234"

bdRaportFeedback='C:\\Dezvoltare\\EXPUR\\NFC\\Report.xlsx'

@app.route('/page/')
def tag():

    return render_template('page.html')

@app.route('/page/' , methods=['POST', 'GET'])
def login():
    if request.method == 'POST':
        ans1 = request.form['ans1']
        ans2 = request.form['ans2']
        ans3 = request.form['ans3']
        ans4 = request.form['ans4']    

        Feedback=openpyxl.load_workbook(bdRaportFeedback, data_only = True)
        rapSheet=Feedback.active
        maxr=rapSheet.max_row
        rapSheet.cell(row=maxr+1, column=2).value=ans1
        rapSheet.cell(row=maxr+1, column=2).value=ans2
        rapSheet.cell(row=maxr+1, column=2).value=ans3
        rapSheet.cell(row=maxr+1, column=2).value=ans4
        Feedback.save(bdRaportFeedback)
    return render_template('finish.html')

app.run(debug="True",host="0.0.0.0", port=3000)
